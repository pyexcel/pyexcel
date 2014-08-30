"""
Design assumption:

It is a MxN formed table

"""

class CSVReader:
    """
    csv reader
    """
    def __init__(self, file):
        import csv
        self.array = []
        reader = csv.reader(open(file, 'rb'), dialect=csv.excel)
        self.array.extend(reader)

    def number_of_rows(self):
        """
        Number of rows in the csv file
        """
        return len(self.array)

    def number_of_columns(self):
        """
        Number of columns in the csv file

        assuming the length of each row is uniform
        """
        if len(self.array) > 1:
            return len(self.array[0])
        else:
            return 0

    def cell_value(self, row, column):
        """
        Random access to the csv cells
        """
        return self.array[row][column]


class XLSReader:
    """
    xls reader
    """
    def __init__(self, file):
        import xlrd
        self.workbook = xlrd.open_workbook(file)
        self.worksheet = self.workbook.sheet_by_index(0)
        
    def number_of_rows(self):
        """
        Number of rows in the xls file
        """
        return self.worksheet.nrows

    def number_of_columns(self):
        """
        Number of columns in the xls file
        """        
        return self.worksheet.ncols

    def cell_value(self, row, column):
        """
        Random access to the xls cells
        """
        return self.worksheet.cell_value(row, column)

class XLSXReader:
    """
    xlsx reader
    """
    def __init__(self, file):
        from openpyxl import load_workbook
        self.wb = load_workbook(file)
        self.ws = self.wb.active
        self.columns = [x for x in 'ABCDEFGHIJKLMNOPQRST']

    def _get_columns(self, index):
        length = len(self.columns)
        if index < length:
            return self.columns[index]
        else:
            return self._get_columns(index/length) + self.columns[index%length]
            
    def number_of_rows(self):
        """
        Number of rows in the xlsx file
        """        
        return self.ws.get_highest_row()

    def number_of_columns(self):
        """
        Number of columns in the xlsx file
        """
        if self.ws.get_highest_row() > 0:
            count = 0
            for i in range(0,self.ws.get_highest_column()):
                if self.cell_value(0, i) is None:
                    break
                count = count + 1
            return count
        else:
            return 0

    def cell_value(self, row, column):
        """
        Random access to the xlsx cells
        """
        actual_row = row + 1
        return self.ws.cell("%s%d" % (self._get_columns(column), actual_row)).value


class Reader:
    """
    Wrapper class to unify csv, xls and xlsx reader
    """
    def __init__(self, file):
        """
        Reader constructor

        Selecting a specific reader according to file extension
        """
        if file.endswith(".xlsx"):
            self.reader = XLSXReader(file)
        elif file.endswith(".xls"):
            self.reader = XLSReader(file)
        elif file.endswith(".csv"):
            self.reader = CSVReader(file)
        else:
            raise NotImplementedError("can not open %s" % file)

    def number_of_rows(self):
        """
        Number of rows in the data file
        """
        return self.reader.number_of_rows()

    def number_of_columns(self):
        """
        Number of columns in the data file
        """
        return self.reader.number_of_columns()

    def cell_value(self, row, column):
        """
        Random access to the data cells
        """
        if row in self.row_range() and column in self.column_range():
            return self.reader.cell_value(row, column)
        else:
            return None

    def row_range(self):
        """
        Utility function to get row range
        """
        return range(0, self.reader.number_of_rows())

    def column_range(self):
        """
        Utility function to get column range
        """
        return range(0, self.reader.number_of_columns())
